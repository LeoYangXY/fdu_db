from django.shortcuts import render

# Create your views here.
from django.utils import timezone

from django.shortcuts import render

# Create your views here.

import openpyxl
from django.shortcuts import render, redirect
from django.http import HttpResponse
from core.models import (Student, Teacher, Course, Enrollment)


def admin_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        data_type = request.POST.get('data_type')

        if not excel_file or not data_type:
            return render(request, 'users/admin_import.html', {'error': '请选择文件和数据类型'})

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            if data_type == 'student':
                success_count = import_students(sheet)
                message = f'成功导入 {success_count} 条学生数据'
            elif data_type == 'teacher':
                success_count = import_teachers(sheet)
                message = f'成功导入 {success_count} 条教师数据'
            elif data_type == 'course':
                success_count = import_courses(sheet)
                message = f'成功导入 {success_count} 条课程数据'
            elif data_type == 'enrollment':  # 新增选课数据导入
                success_count = import_enrollments(sheet)
                message = f'成功导入 {success_count} 条选课数据'
            else:
                message = '未知的数据类型'

            return render(request, 'users/admin_import.html', {'message': message})

        except Exception as e:
            return render(request, 'users/admin_import.html', {'error': f'导入失败: {str(e)}'})

    return render(request, 'users/admin_import.html')



def import_students(sheet):
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 跳过空行
            continue

        # 更新后的字段顺序（移除了account）
        student_id, name, department, major, gender, wechat_openid, password = row[:7]

        # 创建或更新学生记录
        student, created = Student.objects.update_or_create(
            student_id=student_id,
            defaults={
                'name': name,
                'department': department,
                'major': major,
                'gender': gender,
                'wechat_openid': wechat_openid,
                'password': password  # 直接存储明文密码
            }
        )

        count += 1
    return count


def import_teachers(sheet):
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        # 更新后的字段顺序（移除了account）
        teacher_id, name, department, contact, password = row[:5]

        # 创建或更新教师记录
        teacher, created = Teacher.objects.update_or_create(
            teacher_id=teacher_id,
            defaults={
                'name': name,
                'department': department,
                'contact': contact,
                'password': password  # 直接存储明文密码
            }
        )

        count += 1
    return count

def import_courses(sheet):
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        course_code, course_name, department, teacher_id, course_time = row[:5]

        # 检查教师是否存在
        if not Teacher.objects.filter(teacher_id=teacher_id).exists():
            continue

        Course.objects.update_or_create(
            course_code=course_code,
            defaults={
                'course_name': course_name,
                'department': department,
                'teacher_id': teacher_id,
                'course_time': course_time
            }
        )
        count += 1
    return count


# 新增的导入选课数据函数
def import_enrollments(sheet):
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 跳过空行
            continue

        student_id, course_code, semester = row[:3]

        # 检查学生和课程是否存在
        if not Student.objects.filter(student_id=student_id).exists():
            continue
        if not Course.objects.filter(course_code=course_code).exists():
            continue

        Enrollment.objects.update_or_create(
            student_id=student_id,
            course_id=course_code,
            semester=semester,
            defaults={'enrollment_time': timezone.now()}
        )
        count += 1
    return count


def download_template(request, data_type):
    wb = openpyxl.Workbook()
    ws = wb.active

    if data_type == 'student':
        ws.append(['学号', '姓名', '院系', '专业', '性别(M/F)', '微信openid', '密码'])  # 移除了账号
    elif data_type == 'teacher':
        ws.append(['工号', '姓名', '院系', '联系方式', '密码'])  # 移除了账号
    elif data_type == 'course':
        ws.append(['课程代码', '课程名称', '开课院系', '教师工号', '上课时间'])
    elif data_type == 'enrollment':
        ws.append(['学号', '课程代码', '学期(如:2023-秋季)'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={data_type}_template.xlsx'
    wb.save(response)

    return response

def admin_dashboard(request):
    return render(request, 'admin_dashboard.html')
